#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Lancer en CMD "pip install requests" pour installer le module

import json
import time
import re
import requests
import openpyxl
import progressbar
import sys, getopt
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from datetime import datetime

filepath = "./Shiva-RVTools-" + time.strftime('%Y%m%d-%H%M', time.localtime(time.time())) + ".xlsx"
REFRESH_INTERVAL = 200

ERROR = 1
VERBOSE = 2
DEBUG = 3

PROGRESS = False
LOGLEVEL = VERBOSE
MB = 1024 * 1024

opts, args = getopt.getopt(sys.argv[1:], "hpi:s:o:d:", ["pat-id=", "pat-secret=", "output-file="])
for opt, arg in opts:
    if opt == '-h':
        print("Usage : " + sys.argv[0] + " [-h] [-i patid] [-s patsecret] [-o output-file] [-d level] [-p]")
        print("      -h                         : print this help and exit")
        print("      -i PAT-ID                  : API Token PAT ID")
        print("      --pat-id=PAT-ID            : API Token PAT ID")
        print("      -s PAT-SECRET              : API Token PAT SECRET")
        print("      --pat-secret=PAT-SECRET    : API Token PAT SECRET")
        print("      -o PATH-TO-FILE            : Path to output file")
        print("      --output-file=PATH-TO-FILE : API Token PAT SECRET")
        print("      -d LOGLEVEL                : LOG LEVEL (ERROR, VERBOSE, DEBUG)")
        print("      -p                         : Enable Progressbar")
        sys.exit()
    elif opt in ("-i", "--pat-id"):
        pat_id = arg
    elif opt in ("-s", "--pat-secret"):
        pat_secret = arg
    elif opt in ("-o", "--output-file"):
        filepath = arg
    elif opt == "-d":
        LOGLEVEL = eval(arg)
    elif opt == "-p":
        PROGRESS = True

# Fill
blackFill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
progressbar.streams.wrap_stderr()

# URL
base_url = "https://shiva.cloud-temple.com"
auth_url = base_url + "/api/iam/v2/auth/personal_access_token"
tenant_url = base_url + "/api/iam/v2/tenants"
vcenters_url = base_url + "/api/compute/v1/vcenters"
hosts_cluster_url = base_url + "/api/compute/v1/vcenters/host_clusters"
hosts_url = base_url + "/api/compute/v1/vcenters/hosts"
datastore_cluster_url = base_url + "/api/compute/v1/vcenters/datastore_clusters"
datastore_url = base_url + "/api/compute/v1/vcenters/datastores"
vm_url = base_url + "/api/compute/v1/vcenters/virtual_machines"
vm_disks_url = base_url + "/api/compute/v1/vcenters/virtual_disks?virtualMachineId="
network_url = base_url + "/api/compute/v1/vcenters/networks"
vm_controllers_url = base_url + "/api/compute/v1/vcenters/virtual_controllers?virtualMachineId="
vm_nics_url = base_url + "/api/compute/v1/vcenters/network_adapters?virtualMachineId="
vm_snapshot_url = base_url + "/api/compute/v1/vcenters/snapshots?virtualMachineId="


def log(s, level):
    if LOGLEVEL == 1 and level == 1:
        print(s)
    elif LOGLEVEL == 2 and level <= 2:
        print(s)
    elif LOGLEVEL == 3 and level <= 3:
        print(s)


def create_xlsx(filepath):
    workbook = Workbook()
    workbook.save(filepath)
    return workbook


def create_sheet(workbook, name):
    return workbook.create_sheet(title=name)


def rename_current_sheet(sheet, name):
    sheet.title = name


def select_default_sheet(workbook):
    return workbook.active


def select_sheet(workbook, sheetname):
    return workbook[sheetname]


def insert_row(workbook, rowid, num):
    return sheet.insert_row(idx=rowid, amount=num)


def write_array_to_row(sheet, rowid, array):
    for i in range(0, len(array)):
        e = sheet.cell(row=rowid, column=i + 1)
        e.value = str(array[i])


def save_workbook(workbook, filepath):
    workbook.save(filepath)


def init_rvtools():
    # CREATE XLSX
    workbook = create_xlsx(filepath)
    sheet = select_default_sheet(workbook)
    rename_current_sheet(sheet, "vInfo")
    create_sheet(workbook, "vCPU")
    create_sheet(workbook, "vMemory")
    create_sheet(workbook, "vDisk")
    create_sheet(workbook, "vNetwork")
    create_sheet(workbook, "vSnapshot")
    create_sheet(workbook, "vSource")
    create_sheet(workbook, "vCluster")
    create_sheet(workbook, "vHost")
    create_sheet(workbook, "vDatastore")
    create_sheet(workbook, "vHealth")

    vInfoHeaders = [
        "VM",
        "Powerstate",
        "Template",
        "SRM Placeholder",
        "Config status",
        "DNS Name",
        "Connection state",
        "Guest state",
        "Heartbeat",
        "Consolidation Needed",
        "PowerOn",
        "Suspend time",
        "Creation date",
        "Change Version",
        "CPUs",
        "Memory",
        "NICs",
        "Disks",
        "Total disk capacity MiB",
        "min Required EVC Mode Key",
        "Latency Sensitivity",
        "EnableUUID",
        "CBT",
        "Primary IP Address",
        "Network #1",
        "Network #2",
        "Network #3",
        "Network #4",
        "Network #5",
        "Network #6",
        "Network #7",
        "Network #8",
        "Num Monitors",
        "Video Ram KiB",
        "Resource pool",
        "Folder ID",
        "Folder",
        "vApp",
        "DAS protection",
        "FT State",
        "FT Role",
        "FT Latency",
        "FT Bandwidth",
        "FT Sec. Latency",
        "Provisioned MiB",
        "In Use MiB",
        "Unshared MiB",
        "HA Restart Priority",
        "HA Isolation Response",
        "HA VM Monitoring",
        "Cluster rule(s)",
        "Cluster rule name(s)",
        "Boot Required",
        "Boot delay",
        "Boot retry delay",
        "Boot retry enabled",
        "Boot BIOS setup",
        "Reboot PowerOff",
        "EFI Secure boot",
        "Firmware",
        "HW version",
        "HW upgrade status",
        "HW upgrade policy",
        "HW target",
        "Path",
        "Log directory",
        "Snapshot directory",
        "Suspend directory",
        "Annotation",
        "Datacenter",
        "Cluster",
        "Host",
        "OS according to the configuration file",
        "OS according to the VMware Tools",
        "VM ID",
        "SMBIOS UUID",
        "VM UUID",
        "VI SDK Server type",
        "VI SDK API Version",
        "VI SDK Server",
        "VI SDK UUID",
        "isMetroReplicated",
        "isvSphereReplicated"
    ]
    write_array_to_row(select_sheet(workbook, "vInfo"), 1, vInfoHeaders)

    vCPUHeaders = [
        "VM",
        "Powerstate",
        "Template",
        "SRM Placeholder",
        "CPUs",
        "Sockets",
        "Cores p/s",
        "Max",
        "Overall",
        "Level",
        "Shares",
        "Reservation",
        "Entitlement",
        "DRS Entitlement",
        "Limit",
        "Hot Add",
        "Hot Remove",
        "Numa Hotadd Exposed",
        "Annotation",
        "Datacenter",
        "Cluster",
        "Host",
        "Folder",
        "OS according to the configuration file",
        "OS according to the VMware Tools",
        "VM ID",
        "VM UUID",
        "VI SDK Server",
        "VI SDK UUID"
    ]

    write_array_to_row(select_sheet(workbook, "vCPU"), 1, vCPUHeaders)

    vMemoryHeaders = [
        "VM",
        "Powerstate",
        "Template",
        "SRM Placeholder",
        "Size MiB",
        "Memory Reservation Locked To Max",
        "Overhead",
        "Max",
        "Consumed",
        "Consumed Overhead",
        "Private",
        "Shared",
        "Swapped",
        "Ballooned",
        "Active",
        "Entitlement",
        "DRS Entitlement",
        "Level",
        "Shares",
        "Reservation",
        "Limit",
        "Hot Add",
        "Annotation",
        "Datacenter",
        "Cluster",
        "Host",
        "Folder",
        "OS according to the configuration file",
        "OS according to the VMware Tools",
        "VM ID",
        "VM UUID",
        "VI SDK Server",
        "VI SDK UUID"
    ]

    write_array_to_row(select_sheet(workbook, "vMemory"), 1, vMemoryHeaders)

    vDiskHeaders = [
        "VM",
        "Powerstate",
        "Template",
        "SRM Placeholder",
        "Disk",
        "Disk Key",
        "Disk UUID",
        "Disk Path",
        "Capacity MiB",
        "Raw",
        "Disk Mode",
        "Sharing mode",
        "Thin",
        "Eagerly Scrub",
        "Split",
        "Write Through",
        "Level",
        "Shares",
        "Reservation",
        "Limit",
        "Controller",
        "Label",
        "SCSI Unit #",
        "Unit #",
        "Shared Bus",
        "Path",
        "Raw LUN ID",
        "Raw Comp. Mode",
        "Internal Sort Column",
        "Annotation",
        "Datacenter",
        "Cluster",
        "Host",
        "Folder",
        "OS according to the configuration file",
        "OS according to the VMware Tools",
        "VM ID",
        "VM UUID",
        "VI SDK Server",
        "VI SDK UUID"
    ]

    write_array_to_row(select_sheet(workbook, "vDisk"), 1, vDiskHeaders)

    vNetworkHeaders = [
        "VM",
        "Powerstate",
        "Template",
        "SRM Placeholder",
        "NIC label",
        "Adapter",
        "Network",
        "Switch",
        "Connected",
        "Starts Connected",
        "Mac Address",
        "Type",
        "IPv4 Address",
        "IPv6 Address",
        "Direct Path IO",
        "Internal Sort Column",
        "Annotation",
        "Datacenter",
        "Cluster",
        "Host",
        "Folder",
        "OS according to the configuration file",
        "OS according to the VMware Tools",
        "VM ID",
        "VM UUID",
        "VI SDK Server",
        "VI SDK UUID"
    ]

    write_array_to_row(select_sheet(workbook, "vNetwork"), 1, vNetworkHeaders)

    vSnapshotHeaders = [
        "VM",
        "Powerstate",
        "Name",
        "Description",
        "Date / time",
        "Filename",
        "Size MiB (vmsn)",
        "Size MiB (total)",
        "Quiesced",
        "State",
        "Annotation",
        "Datacenter",
        "Cluster",
        "Host",
        "Folder",
        "OS according to the configuration file",
        "OS according to the VMware Tools",
        "VM ID",
        "VM UUID",
        "VI SDK Server",
        "VI SDK UUID"
    ]

    write_array_to_row(select_sheet(workbook, "vSnapshot"), 1, vSnapshotHeaders)

    vSourceHeaders = [
        "Name",
        "OS type",
        "API type",
        "API version",
        "Version",
        "Patch level",
        "Build",
        "Fullname",
        "Product name",
        "Product version",
        "Product line",
        "Vendor",
        "VI SDK Server",
        "VI SDK UUID"
    ]

    write_array_to_row(select_sheet(workbook, "vSource"), 1, vSourceHeaders)

    vDatastoresHeaders = [
        "Name",
        "Config status",
        "Address",
        "Accessible",
        "Type",
        "# VMs total",
        "# VMs",
        "Capacity MiB",
        "Provisioned MiB",
        "In Use MiB",
        "Free MiB",
        "Free %",
        "SIOC enabled",
        "SIOC Threshold",
        "# Hosts",
        "Hosts",
        "Cluster name",
        "Cluster capacity MiB",
        "Cluster free space MiB",
        "Block size",
        "Max Blocks",
        "# Extents",
        "Major Version",
        "Version",
        "VMFS Upgradeable",
        "MHA",
        "URL",
        "Object ID",
        "VI SDK Server",
        "VI SDK UUID"
    ]

    write_array_to_row(select_sheet(workbook, "vDatastore"), 1, vDatastoresHeaders)

    vHostClusterHeaders = [
        "Name",
        "Config status",
        "OverallStatus",
        "NumHosts",
        "numEffectiveHosts",
        "TotalCpu",
        "NumCpuCores",
        "NumCpuThreads",
        "Effective Cpu",
        "TotalMemory",
        "Effective Memory",
        "Num VMotions",
        "HA enabled",
        "Failover Level",
        "AdmissionControlEnabled",
        "Host monitoring",
        "HB Datastore Candidate Policy",
        "Isolation Response",
        "Restart Priority",
        "Cluster Settings",
        "Max Failures",
        "Max Failure Window",
        "Failure Interval",
        "Min Up Time",
        "VM Monitoring",
        "DRS enabled",
        "DRS default VM behavior",
        "DRS vmotion rate",
        "DPM enabled",
        "DPM default behavior",
        "DPM Host Power Action Rate",
        "Object ID",
        "VI SDK Server",
        "VI SDK UUID"
    ]

    write_array_to_row(select_sheet(workbook, "vCluster"), 1, vHostClusterHeaders)

    vHostHeaders = [
        "Host",
        "Datacenter",
        "Cluster",
        "Config status",
        "in Maintenance Mode",
        "in Quarantine Mode",
        "vSAN Fault Domain Name",
        "CPU Model",
        "Speed",
        "HT Available",
        "HT Active",
        "# CPU",
        "Cores per CPU",
        "# Cores",
        "CPU usage %",
        "# Memory",
        "Memory usage %",
        "Console",
        "# NICs",
        "# HBAs",
        "# VMs total",
        "# VMs",
        "VMs per Core",
        "# vCPUs",
        "vCPUs per Core",
        "vRAM",
        "VM Used memory",
        "VM Memory Swapped",
        "VM Memory Ballooned",
        "VMotion support",
        "Storage VMotion support",
        "Current EVC",
        "Max EVC",
        "Assigned License(s)",
        "ATS Heartbeat",
        "ATS Locking",
        "Current CPU power man. policy",
        "Supported CPU power man.",
        "Host Power Policy",
        "ESX Version",
        "Boot time",
        "DNS Servers",
        "DHCP",
        "Domain",
        "DNS Search Order",
        "NTP Server(s)",
        "NTPD running",
        "Time Zone",
        "Time Zone Name",
        "GMT Offset",
        "Vendor",
        "Model",
        "Serial number",
        "Service tag",
        "OEM specific string",
        "BIOS Vendor",
        "BIOS Version",
        "BIOS Date",
        "Object ID",
        "AutoDeploy.MachineIdentity",
        "UUID",
        "VI SDK Server",
        "VI SDK UUID"
    ]

    write_array_to_row(select_sheet(workbook, "vHost"), 1, vHostHeaders)

    vHealthHeader = [
        "Name",
        "Message",
        "Message type",
        "VI SDK Server",
        "VI SDK UUID"
    ]

    write_array_to_row(select_sheet(workbook, "vHealth"), 1, vHealthHeader)

    return workbook


def refresh_token(headers_auth):
    request_auth = requests.post(auth_url, data=headers_auth)
    token = request_auth.text

    if request_auth.status_code == 200:
        log("-> Authentification OK", 3)
    else:
        log("-> Erreur d'authentification", 1)
        exit()
    headers = {'Authorization': 'Bearer ' + token, 'accept': 'application/json'}
    return headers


# AUTHENTIFICATION
log("\nAUTHENTIFICATION SHIVA", 2)
headers_auth = {"id": pat_id, "secret": pat_secret}
request_auth = requests.post(auth_url, data=headers_auth)
token = request_auth.text

if request_auth.status_code == 200:
    log("-> Authentification OK", 3)
else:
    log("-> Erreur d'authentification", 1)
    exit()

# REQUEST TENANT ID
log("\nREQUEST 'TENANT ID'", 2)

headers = {'Authorization': 'Bearer ' + token, 'accept': 'application/json'}
request_ten = requests.get(tenant_url, headers=headers)
if request_ten.status_code == 200:
    json_ten = json.loads(request_ten.text)
    for data_ten in json_ten:
        log("-> Nom Tenant : " + data_ten['name'] + " \n-> ID Tenant : " + data_ten['id'] + " \n-> Company ID : " +
            data_ten['companyId'], 3)
else:
    log("-> Erreur sur la requête TENANT ID", 1)
    exit()

# CREATE WORKBOOK
workbook = init_rvtools()

# REQUEST VIRTUAL DATACENTERS
log("\nREQUEST 'VIRTUAL DATACENTERS'", 2)
sourcerowid = 2
vcenterlist = {}
request_vcenters = requests.get(vcenters_url, headers=headers)
if request_vcenters.status_code == 200:
    json_vc = json.loads(request_vcenters.text)
    for vc in json_vc:
        vSource = [
            vc.get('name', ''),  # Name
            vc.get('osType', ''),  # OS type
            vc.get('apiType', ''),  # API type
            vc.get('apiVersion', ''),  # API version
            vc.get('version', ''),  # Version
            vc.get('patchLevel', ''),  # Patch level
            vc.get('build', ''),  # Build
            vc.get('fullName', ''),  # Fullname
            vc.get('licenseProductName', ''),  # Product name
            vc.get('licenseProductVersion', ''),  # Product version
            vc.get('licenseProductLine', ''),  # Product line
            vc.get('vendor', ''),  # Vendor
            "Shiva API",  # VI SDK Server
            vc.get('id', ''),  # VI SDK UUID
        ]
        vcenterlist[vc.get('id', '')] = vc.get('name', '')
        write_array_to_row(select_sheet(workbook, "vSource"), sourcerowid, vSource)
        sourcerowid = sourcerowid + 1
    save_workbook(workbook, filepath)

# REQUEST HOSTS CLUSTERS
log("\nREQUEST 'HOSTS CLUSTERS'", 2)
hostclusterrowid = 2
host2Cluster = {}
request_hostsclusters = requests.get(hosts_cluster_url, headers=headers)
if request_hostsclusters.status_code == 200:
    json_hostclusters = json.loads(request_hostsclusters.text)

    for vc in json_hostclusters:

        request_hostscluster = requests.get(hosts_cluster_url + "/" + vc.get('id', ''), headers=headers)
        if request_hostscluster.status_code == 200:
            hc = json.loads(request_hostscluster.text)
            if vc.get('machineManagerId', '') not in host2Cluster.keys():
                host2Cluster[vc.get('machineManagerId', '')] = {}

            for host in hc.get('hosts', []):
                host2Cluster[vc.get('machineManagerId', '')][host.get('id', '')] = {}
                host2Cluster[vc.get('machineManagerId', '')][host.get('id', '')]['clusterName'] = hc.get('name', '')
                host2Cluster[vc.get('machineManagerId', '')][host.get('id', '')]['datacenterName'] = hc.get(
                    'datacenterName', '')
            vCluster = [
                hc.get('name', ''),  # Name
                "",  # Config status
                "",  # OverallStatus
                len(hc.get('hosts', [])),  # NumHosts
                len(hc.get('hosts', [])),  # numEffectiveHosts
                hc.get('metrics', {})['totalCpu'],  # TotalCpu
                "",  # NumCpuCores
                "",  # NumCpuThreads
                "",  # Effective Cpu
                hc.get('metrics', {})['totalMemory'],  # TotalMemory
                "",  # Effective Memory
                "",  # Num VMotions
                "",  # HA enabled
                "",  # Failover Level
                "",  # AdmissionControlEnabled
                "",  # Host monitoring
                "",  # HB Datastore Candidate Policy
                "",  # Isolation Response
                "",  # Restart Priority
                "",  # Cluster Settings
                "",  # Max Failures
                "",  # Max Failure Window
                "",  # Failure Interval
                "",  # Min Up Time
                "",  # VM Monitoring
                hc.get('drsConfig', {})['enabled'],  # DRS enabled
                hc.get('drsConfig', {})['automationLevel'],  # DRS default VM behavior
                hc.get('drsConfig', {})['migrationThreshold'],  # DRS vmotion rate
                "",  # DPM enabled
                "",  # DPM default behavior
                "",  # DPM Host Power Action Rate
                hc.get('moref', ''),  # Object ID
                vcenterlist[vc.get('machineManagerId', '')],  # VI SDK Server
                vc.get('machineManagerId', ''),  # VI SDK UUID
            ]

            write_array_to_row(select_sheet(workbook, "vCluster"), hostclusterrowid, vCluster)
            hostclusterrowid = hostclusterrowid + 1
        else:
            log('Error getting Host Clusters info', 1)
    save_workbook(workbook, filepath)
else:
    log('Error getting Host Clusters info', 1)

# REQUEST HOSTS
hostrowid = 2
request_hosts = requests.get(hosts_url, headers=headers)
if request_hosts.status_code == 200:
    json_hosts = json.loads(request_hosts.text)
    for host in json_hosts:
        vHost = [
            host.get('name', ''),  # Host
            host2Cluster[host.get('machineManagerId', '')][host.get('moref', '')]['datacenterName'],  # Datacenter
            host2Cluster[host.get('machineManagerId', '')][host.get('moref', '')]['clusterName'],  # Cluster
            "",  # Config status
            host.get('maintenanceStatus', ''),  # in Maintenance Mode
            "",  # in Quarantine Mode
            "",  # vSAN Fault Domain Name
            "",  # CPU Model
            host.get('metrics', {})['cpu']['cpuMhz'],  # Speed
            "",  # HT Available
            "",  # HT Active
            "",  # # CPU
            "",  # Cores per CPU
            host.get('metrics', {})['cpu']['cpuCores'],  # # Cores
            int((host.get('metrics', {})['cpu']['overallCpuUsage'] / (
                        host.get('metrics', {})['cpu']['cpuMhz'] * host.get('metrics', {})['cpu']['cpuCores'])) * 100),
            # CPU usage %
            int(host.get('metrics', {})['memory']['memorySize'] / MB),  # # Memory
            int((int(host.get('metrics', {})['memory']['memoryUsage']) / int(
                host.get('metrics', {})['memory']['memorySize'] / MB)) * 100),  # Memory usage %
            "",  # Console
            "",  # # NICs
            "",  # # HBAs
            len(host.get('virtualMachines', [])),  # # VMs total
            len(host.get('virtualMachines', [])),  # # VMs
            len(host.get('virtualMachines', [])) / host.get('metrics', {})['cpu']['cpuCores'],  # VMs per Core
            "",  # # vCPUs
            "",  # vCPUs per Core
            "",  # vRAM
            "",  # VM Used memory
            "",  # VM Memory Swapped
            "",  # VM Memory Ballooned
            "",  # VMotion support
            "",  # Storage VMotion support
            "",  # Current EVC
            "",  # Max EVC
            "",  # Assigned License(s)
            "",  # ATS Heartbeat
            "",  # ATS Locking
            "",  # Current CPU power man. policy
            "",  # Supported CPU power man.
            "",  # Host Power Policy
            host.get('metrics', {})['esx']['fullName'],  # ESX Version
            time.strftime('%d/%m/%Y %H:%M:%S', time.localtime(time.time() - host.get('metrics', {})['uptime'])),
            # Boot time
            "",  # DNS Servers
            "",  # DHCP
            "",  # Domain
            "",  # DNS Search Order
            "",  # NTP Server(s)
            "",  # NTPD running
            "",  # Time Zone
            "",  # Time Zone Name
            "",  # GMT Offset
            "",  # Vendor
            "",  # Model
            "",  # Serial number
            "",  # Service tag
            "",  # OEM specific string
            "",  # BIOS Vendor
            "",  # BIOS Version
            "",  # BIOS Date
            host.get('moref', ''),  # Object ID
            "",  # AutoDeploy.MachineIdentity
            host.get('id', ''),  # UUID
            vcenterlist[host.get('machineManagerId', '')],  # VI SDK Server
            host.get('machineManagerId', ''),  # VI SDK UUID
        ]

        write_array_to_row(select_sheet(workbook, "vHost"), hostrowid, vHost)
        hostrowid = hostrowid + 1
    save_workbook(workbook, filepath)
else:
    print('Error getting Host info')

# REQUEST DATASTORE CLUSTERS
log("\nREQUEST 'DATASTORE CLUSTERS' and DATASTORES", 2)
datastorerowid = 2

request_vdsclusters = requests.get(datastore_cluster_url, headers=headers)
if request_vdsclusters.status_code == 200:
    json_vds = json.loads(request_vdsclusters.text)
    datastore2vds = {}
    for vds in json_vds:
        datastores = vds.get('datastores', [])
        metrics = vds.get('metrics', {})
        vd = {}
        vd['vdsname'] = vds.get('name', '')
        vd['capacity'] = int(metrics.get('maxCapacity', '') / MB)
        vd['freeCapacity'] = int(metrics.get('freeCapacity', '') / MB)
        vd['vcenterId'] = vds.get('machineManagerId')
        for datastore in datastores:
            datastore2vds[datastore] = {}
            datastore2vds[datastore] = vd

request_datastores = requests.get(datastore_url, headers=headers)
if request_datastores.status_code == 200:
    json_ds = json.loads(request_datastores.text)
    for ds in json_ds:
        vDatastore = [
            ds.get('name', ''),  # Name
            "",  # Config status
            "",  # Address
            bool(ds.get('accessible', '')),  # Accessible
            ds.get('type', ''),  # Type
            "",  # # VMs total
            ds.get('virtualMachinesNumber', ''),  # # VMs
            int(ds.get('maxCapacity', '') / MB),  # Capacity MiB
            "",  # Provisioned MiB
            int(ds.get('maxCapacity', 0) / MB) - int(ds.get('freeCapacity', '') / MB),  # In Use MiB
            int(ds.get('freeCapacity', '') / MB),  # Free MiB
            int((int(ds.get('freeCapacity', '') / MB) / int(ds.get('maxCapacity', '') / MB)) * 100),  # Free %
            "",  # SIOC enabled
            "",  # SIOC Threshold
            ds.get('hostsNumber', ''),  # # Hosts
            ', '.join(ds.get('hostsNames', [])),  # Hosts
            datastore2vds[ds.get('id', '')]['vdsname'],  # Cluster name
            datastore2vds[ds.get('id', '')]['capacity'],  # Cluster capacity MiB
            datastore2vds[ds.get('id', '')]['freeCapacity'],  # Cluster free space MiB
            "",  # Block size
            "",  # Max Blocks
            "",  # # Extents
            "",  # Major Version
            "",  # Version
            "",  # VMFS Upgradeable
            "",  # MHA
            "ds:///vmfs/volumes/" + ds.get('id', '') + "/",  # URL
            ds.get('moref', ''),  # Object ID
            vcenterlist[datastore2vds[ds.get('id', '')]['vcenterId']],
            datastore2vds[ds.get('id', '')]['vcenterId'],  # VI SDK UUID
        ]

        write_array_to_row(select_sheet(workbook, "vDatastore"), datastorerowid, vDatastore)
        datastorerowid = datastorerowid + 1
    save_workbook(workbook, filepath)

# REQUEST VIRTUAL NETWORKS
log("\nREQUEST 'VIRTUAL NETWORKS'", 2)

networklist = {}
request_networks = requests.get(network_url, headers=headers)
if request_networks.status_code == 200:
    for network in json.loads(request_networks.text):
        networklist[network.get('id', '')] = {}
        networklist[network.get('id', '')]['name'] = network.get('name', '')
        networklist[network.get('id', '')]['switchName'] = network.get('switchName', '')

# REQUEST VIRTUAL MACHINES
log("\nREQUEST 'VIRTUAL MACHINES'", 2)

request_vm = requests.get(vm_url, headers=headers)
if request_vm.status_code == 200:
    json_vm = json.loads(request_vm.text)
    log("\nVirtual machines list :", 2)
    rowid = 1
    diskrowid = 2
    nicrowid = 2
    snaprowid = 2
    vhealthrowid = 2
    bar = progressbar.ProgressBar(max_value=len(json_vm), redirect_stdout=True) if PROGRESS else ""
    for data_vm in json_vm:
        log("  - VM : " + data_vm['name'], 2)
        rowid = rowid + 1
        if rowid % REFRESH_INTERVAL == 0:
            headers = refresh_token(headers_auth)
        vm_id = data_vm['id']
        vm_info_url = vm_url + "/" + vm_id
        vm_disk_info_url = vm_disks_url + vm_id
        vm_controller_info_url = vm_controllers_url + vm_id
        vm_nics_info_url = vm_nics_url + vm_id
        vm_snapshots_info_url = vm_snapshot_url + vm_id
        request_vm_info = requests.get(vm_info_url, headers=headers)
        if request_vm_info.status_code == 200:
            json_vm_info = json.loads(request_vm_info.text)

            # Parse controller info
            request_vm_controllers_info = requests.get(vm_controller_info_url, headers=headers)
            controllers = {}
            if request_vm_controllers_info.status_code == 200:
                json_vm_controllers_info = json.loads(request_vm_controllers_info.text)
                for controller in json_vm_controllers_info:
                    controllers[controller.get('id', '')] = {}
                    controllers[controller.get('id', '')]['sharedBus'] = controller.get('sharedBus', '')
                    controllers[controller.get('id', '')]['label'] = controller.get('label', '')
                    controllers[controller.get('id', '')]['controller'] = controller.get('summary', '')
            else:
                log(f"Error getting controllers info for VM {vm_id}", 1)

            # Parse NICS info
            nicsInfo = []
            request_vm_nics_info = requests.get(vm_nics_info_url, headers=headers)
            if request_vm_nics_info.status_code == 200:
                json_vm_nics_info = json.loads(request_vm_nics_info.text)
                for nic in json_vm_nics_info:
                    nicInfo = {}
                    nicInfo['id'] = nic.get('id', '')
                    nicInfo['name'] = nic.get('name', '')
                    nicInfo['type'] = nic.get('type', '')
                    nicInfo['macType'] = nic.get('macType', '')
                    nicInfo['macAddress'] = nic.get('macAddress', '')
                    nicInfo['connected'] = nic.get('connected', '')
                    nicInfo['autoConnect'] = nic.get('autoConnect', '')
                    nicInfo['networkId'] = nic.get('networkId', '')
                    nicsInfo.append(nicInfo)
            else:
                log(f"Error getting nics info for VM {vm_id}", 1)

            # Parse disk info
            disksInfo = []
            request_vm_disks_info = requests.get(vm_disk_info_url, headers=headers)
            if request_vm_disks_info.status_code == 200:
                json_vm_disks_info = json.loads(request_vm_disks_info.text)
                isMetroReplicated = True
                total_disks_capacity = 0
                for disk in json_vm_disks_info:
                    total_disks_capacity = total_disks_capacity + disk.get('capacity')
                    diskinfo = {}
                    diskinfo["disk"] = disk.get('name', '')
                    diskinfo["capacity"] = int(disk.get('capacity', '') / MB)
                    diskinfo["UUID"] = disk.get('nativeId', '')
                    diskinfo["raw"] = False
                    diskinfo["mode"] = disk.get('diskMode', '')
                    diskinfo["thin"] = True if disk.get('provisioningType', '') == "dynamic" else False
                    diskinfo["unit#"] = disk.get('diskUnitNumber', '')
                    diskinfo["path"] = disk.get('diskPath', '')
                    diskinfo["datastore"] = disk.get('datastoreName', '')
                    diskinfo["sharedBus"] = controllers[disk.get('controllerId', '')]['sharedBus'] if disk.get(
                        'controllerId', '') in controllers else ""
                    diskinfo["label"] = controllers[disk.get('controllerId', '')]['label'] if disk.get('controllerId',
                                                                                                       '') in controllers else ""
                    diskinfo["controller"] = controllers[disk.get('controllerId', '')]['controller'] if disk.get(
                        'controllerId', '') in controllers else ""
                    if not "_mm" in diskinfo["datastore"]:
                        isMetroReplicated = False
                    disksInfo.append(diskinfo)
            else:
                log(f"Error getting disks info for VM {vm_id}", 1)

            # Parse snapshots info
            snapsInfo = []
            request_vm_snaps_info = requests.get(vm_snapshots_info_url, headers=headers)
            if request_vm_snaps_info.status_code == 200:
                json_vm_snap_info = json.loads(request_vm_snaps_info.text)
                for snap in json_vm_snap_info:
                    snapInfo = {}
                    snapInfo['id'] = snap.get('id', '')
                    snapInfo['name'] = snap.get('name', '')
                    snapInfo['description'] = snap.get('description', '')
                    snapInfo['createTime'] = snap.get('createTime', '')
                    snapInfo['quiesced'] = snap.get('quiesced', '')
                    snapsInfo.append(snapInfo)
            else:
                log(f"Error getting snapshots info for VM {vm_id}", 1)

            # Parse Portgroups
            vmnetworks = []
            for pg in json_vm_info.get('distributedVirtualPortGroupIds', ''):
                vmnetworks.append(networklist[pg]['name'])

            # Parse ExtraConfig and find vmtools OS
            vmtoolOS = ""
            extraconfig = json_vm_info.get('extraConfig', '')
            for config in extraconfig:
                if config["key"] == "guestOS.detailed.data":
                    vmtoolOSConfig = config["value"].split('=')
                    vmtoolOS = vmtoolOSConfig[len(vmtoolOSConfig) - 1].replace('\'', '')

            primaryIp = ""
            try:
                primaryIp = json_vm_info.get('ipAddresses', {}).get('primary', '')
            except IndexError:
                primaryIp = ""

            isvSphereReplicated = 'replicationConfig' in data_vm

            vInfo = [
                data_vm['name'],  # VM
                "poweredOn" if json_vm_info.get('powerState', '') == "running" else "poweredOff" if json_vm_info.get(
                    'powerState', '') == "stopped" else json_vm_info.get('powerState', ''),  # Powerstate
                json_vm_info.get('template', ''),  # Template
                "",  # SRM Placeholder
                "",  # Config status
                json_vm_info.get('dnsName', ''),  # DNS Name
                json_vm_info.get('status', ''),  # Connection State
                "",  # Guest state
                "",  # Heartbeat
                json_vm_info.get('consolidationNeeded', ''),  # Consolidation Needed
                "",  # PowerOn
                "",  # Suspend time
                "",  # Creation date
                "",  # Change Version
                json_vm_info.get('cpu', ''),  # CPUs
                int(json_vm_info.get('memory', '') / MB),  # Memory
                len(json_vm_info.get('ipAddresses', {}).get('statics', '')),  # NICs
                len(json_vm_disks_info),  # Disks
                int(total_disks_capacity / MB),  # Total disk capacity MiB
                "",  # min Required EVC Mode Key
                "",  # Latency Sensitivity
                "",  # EnableUUID
                "",  # CBT
                primaryIp,  # Primary IP Address
                vmnetworks[0] if len(vmnetworks) > 0 else "",  # Network #1
                vmnetworks[1] if len(vmnetworks) > 1 else "",  # Network #2
                vmnetworks[2] if len(vmnetworks) > 2 else "",  # Network #3
                vmnetworks[3] if len(vmnetworks) > 3 else "",  # Network #4
                vmnetworks[4] if len(vmnetworks) > 4 else "",  # Network #5
                vmnetworks[5] if len(vmnetworks) > 5 else "",  # Network #6
                vmnetworks[6] if len(vmnetworks) > 6 else "",  # Network #7
                vmnetworks[7] if len(vmnetworks) > 7 else "",  # Network #8
                1,  # Num Monitors
                "",  # Video Ram KiB
                "",  # Resource pool
                "",  # Folder ID
                "/",  # Folder
                "",  # vApp
                "",  # DAS protection
                "",  # FT State
                "",  # FT Role
                "",  # FT Latency
                "",  # FT Bandwidth
                "",  # FT Sec. Latency
                int((json_vm_info.get('storage', {}).get('committed') + json_vm_info.get('storage', {}).get(
                    'uncommitted')) / MB),  # Provisioned MiB
                int(json_vm_info.get('storage', {}).get('committed') / MB),  # In Use MiB
                "",  # Unshared MiB
                "",  # HA Restart Priority
                "",  # HA Isolation Response
                "",  # HA VM Monitoring
                "",  # Cluster rule(s)
                "",  # Cluster rule name(s)
                "",  # Boot Required
                json_vm_info.get('bootOptions', {}).get('bootDelay'),  # Boot delay
                json_vm_info.get('bootOptions', {}).get('bootRetryDelay'),  # Boot retry delay
                json_vm_info.get('bootOptions', {}).get('bootRetryEnabled'),  # Boot retry enabled
                json_vm_info.get('bootOptions', {}).get('enterBIOSSetup'),  # Boot BIOS Setup
                "",  # Reboot PowerOff
                json_vm_info.get('bootOptions', {}).get('efiSecureBootEnabled'),  # EFI Secure boot -
                json_vm_info.get('bootOptions', {}).get('firmware'),  # Firmware
                json_vm_info.get('hardwareVersion', ''),  # HW Version
                "",  # HW Upgrade status
                "",  # HW Upgrade policy
                "",  # HW Target
                "",  # Path
                "",  # Log directory
                "",  # Snapshot directory
                "",  # Suspend directory
                "",  # Annotation
                json_vm_info.get('datacenterName', ''),  # Datacenter
                json_vm_info.get('hostClusterName', ''),  # Cluster
                json_vm_info.get('hostName', ''),  # Host
                json_vm_info.get('operatingSystemName', ''),  # OS According to the configuration file
                vmtoolOS,  # OS according to the VMware Tools
                json_vm_info.get('moref', ''),  # VM ID -------------
                "",  # SMBIOS UUID -------------
                json_vm_info.get('id', ''),  # VM UUID ------------
                "Shiva API",  # VI SDK Server type
                "Shiva API",  # VI SDK API Version
                json_vm_info.get('machineManagerName', ''),  # VI SDK Server
                json_vm_info.get('machineManagerId', ''),  # VI SDK UUID
                isMetroReplicated,  # isMetroReplicated
                isvSphereReplicated
            ]

            write_array_to_row(select_sheet(workbook, "vInfo"), rowid, vInfo)

            if json_vm_info.get('tools', '') in ["toolsNotRunning", "toolsOld", "toolsNotInstalled"]:
                toolsHeath = [
                    data_vm['name'],  # VM
                    "VMware tools are out of date, not running or not installed!",
                    "VM Tools",
                    json_vm_info.get('machineManagerName', ''),  # VI SDK Server
                    json_vm_info.get('machineManagerId', ''),  # VI SDK UUID
                ]
                write_array_to_row(select_sheet(workbook, "vHealth"), vhealthrowid, toolsHeath)
                vhealthrowid = vhealthrowid + 1

            vCPU = [
                data_vm['name'],  # VM
                json_vm_info.get('powerState', ''),  # Powerstate
                json_vm_info.get('template', ''),  # Template
                "",  # SRM Placeholder
                json_vm_info.get('cpu', ''),  # CPUs
                json_vm_info.get('cpu', '') / json_vm_info.get('numCoresPerSocket', ''),  # Sockets
                json_vm_info.get('numCoresPerSocket', ''),  # Cores p/s
                "",  # Max
                "",  # Overall
                "",  # Level
                "",  # Shares
                "",  # Reservation
                "",  # Entitlement
                "",  # DRS Entitlement
                -1,  # Limit
                json_vm_info.get('cpuHotAddEnabled', ''),  # Hot Add
                json_vm_info.get('cpuHotRemoveEnabled', ''),  # Hot Remove
                "",  # Numa Hotadd Exposed
                "",  # Annotation
                json_vm_info.get('datacenterName', ''),  # Datacenter
                json_vm_info.get('hostClusterName', ''),  # Cluster
                json_vm_info.get('hostName', ''),  # Host
                "",  # Folder
                json_vm_info.get('operatingSystemName', ''),  # OS According to the configuration file
                vmtoolOS,  # OS according to the VMware Tools
                json_vm_info.get('moref', ''),  # VM ID
                json_vm_info.get('id', ''),  # VM UUID
                json_vm_info.get('machineManagerName', ''),  # VI SDK Server
                json_vm_info.get('machineManagerId', ''),  # VI SDK UUID
            ]

            write_array_to_row(select_sheet(workbook, "vCPU"), rowid, vCPU)

            vMemory = [
                data_vm['name'],  # VM
                json_vm_info.get('powerState', ''),  # Powerstate
                json_vm_info.get('template', ''),  # Template
                "",  # SRM Placeholder
                int(json_vm_info.get('memory', '') / MB),  # Size MiB
                "",  # Memory Reservation Locked To Max
                "",  # Overhead
                "",  # Max
                "",  # Consumed
                "",  # Consumed Overhead
                "",  # Private
                "",  # Shared
                int(json_vm_info.get('swappedMemory', '') / MB),  # Swapped
                int(json_vm_info.get('balloonedMemory', '') / MB),  # Ballooned
                "",  # Active
                "",  # Entitlement
                "",  # DRS Entitlement
                "",  # Level
                "",  # Shares
                "",  # Reservation
                "",  # Limit
                json_vm_info.get('memoryHotAddEnabled', ''),  # Hot Add
                "",  # Annotation
                json_vm_info.get('datacenterName', ''),  # Datacenter
                json_vm_info.get('hostClusterName', ''),  # Cluster
                json_vm_info.get('hostName', ''),  # Host
                "",  # Folder
                json_vm_info.get('operatingSystemName', ''),  # OS According to the configuration file
                vmtoolOS,  # OS according to the VMware Tools
                json_vm_info.get('moref', ''),  # VM ID
                json_vm_info.get('id', ''),  # VM UUID
                json_vm_info.get('machineManagerName', ''),  # VI SDK Server
                json_vm_info.get('machineManagerId', ''),  # VI SDK UUID
            ]

            write_array_to_row(select_sheet(workbook, "vMemory"), rowid, vMemory)

            # json_vm_info.get('toolsVersion', '')

            # We only take care of first disk... not perfect but OK
            if len(disksInfo) > 0:
                myarray = disksInfo[0]['path'].split(" ")
                folder = ''.join(myarray[1::]).split("/")[0]

                if folder != data_vm['name']:
                    vmname = data_vm['name']
                    myarray = disksInfo[0]['path'].split(" ")
                    folder = ''.join(myarray[1::]).split("/")[0]
                    datastore = myarray[0].split("[")[1].split("]")[0]
                    inconsistentFolder = [
                        vmname,
                        f"Inconsistent Foldername! VMname = {vmname} Foldername = {folder} Datastore = {datastore}",
                        "Foldername",
                        json_vm_info.get('machineManagerName', ''),  # VI SDK Server
                        json_vm_info.get('machineManagerId', '')  # VI SDK UUID
                    ]
                    write_array_to_row(select_sheet(workbook, "vHealth"), vhealthrowid, inconsistentFolder)
                    vhealthrowid = vhealthrowid + 1

            for disk in disksInfo:
                vDisk = [
                    data_vm['name'],  # VM
                    json_vm_info.get('powerState', ''),  # Powerstate
                    json_vm_info.get('template', ''),  # Template
                    "",  # SRM Placeholder
                    disk["disk"],  # Disk
                    "",  # Disk Key
                    disk["UUID"],  # Disk UUID
                    "",  # Disk Path
                    disk["capacity"],  # Capacity MiB
                    disk["raw"],  # Raw
                    disk["mode"],  # Disk Mode
                    "",  # Sharing mode
                    disk["thin"],  # Thin
                    "",  # Eagerly Scrub
                    "",  # Split
                    "",  # Write Through
                    "",  # Level
                    "",  # Shares
                    "",  # Reservation
                    -1,  # Limit
                    disk["controller"],  # Controller
                    disk["label"],  # Label
                    "",  # SCSI Unit #
                    disk["unit#"],  # Unit #
                    disk["sharedBus"],  # Shared Bus
                    disk["path"],  # Path
                    "",  # Raw LUN ID
                    "",  # Raw Comp. Mode
                    "",  # Internal Sort Column
                    "",  # Annotation
                    json_vm_info.get('datacenterName', ''),  # Datacenter
                    json_vm_info.get('hostClusterName', ''),  # Cluster
                    json_vm_info.get('hostName', ''),  # Host
                    "",  # Folder
                    json_vm_info.get('operatingSystemName', ''),  # OS According to the configuration file
                    vmtoolOS,  # OS according to the VMware Tools
                    json_vm_info.get('moref', ''),  # VM ID
                    json_vm_info.get('id', ''),  # VM UUID
                    json_vm_info.get('machineManagerName', ''),  # VI SDK Server
                    json_vm_info.get('machineManagerId', '')  # VI SDK UUID
                ]

                write_array_to_row(select_sheet(workbook, "vDisk"), diskrowid, vDisk)

                diskrowid = diskrowid + 1

            for nic in nicsInfo:
                vNics = [
                    data_vm['name'],  # VM
                    json_vm_info.get('powerState', ''),  # Powerstate
                    json_vm_info.get('template', ''),  # Template
                    "",  # SRM Placeholder
                    nic['name'],  # NIC label
                    nic['type'],  # Adapter
                    networklist[nic['networkId']]['name'] if len(nic['networkId']) > 0 else "",  # Network
                    networklist[nic['networkId']]['switchName'] if len(nic['networkId']) > 0 else "",  # Switch
                    nic['connected'],  # Connected
                    nic['autoConnect'],  # Starts Connected
                    nic['macAddress'],  # Mac Address
                    nic['macType'],  # Type
                    "",  # IPv4 Address
                    "",  # IPv6 Address
                    "",  # Direct Path IO
                    "",  # Internal Sort Column
                    "",  # Annotation
                    json_vm_info.get('datacenterName', ''),  # Datacenter
                    json_vm_info.get('hostClusterName', ''),  # Cluster
                    json_vm_info.get('hostName', ''),  # Host
                    "",  # Folder
                    json_vm_info.get('operatingSystemName', ''),  # OS According to the configuration file
                    vmtoolOS,  # OS according to the VMware Tools
                    json_vm_info.get('moref', ''),  # VM ID
                    json_vm_info.get('id', ''),  # VM UUID
                    json_vm_info.get('machineManagerName', ''),  # VI SDK Server
                    json_vm_info.get('machineManagerId', ''),  # VI SDK UUID
                ]

                write_array_to_row(select_sheet(workbook, "vNetwork"), nicrowid, vNics)
                nicrowid = nicrowid + 1

            for snap in snapsInfo:
                vSnap = [
                    data_vm['name'],  # VM
                    json_vm_info.get('powerState', ''),  # Powerstate
                    snap['name'],  # Name
                    snap['description'],  # Description
                    datetime.fromtimestamp(int(snap['createTime'] / 1000)),  # Date / time
                    "",  # Filename
                    "",  # Size MiB (vmsn)
                    "",  # Size MiB (total)
                    snap['quiesced'],  # Quiesced
                    "",  # State
                    "",  # Annotation
                    json_vm_info.get('datacenterName', ''),  # Datacenter
                    json_vm_info.get('hostClusterName', ''),  # Cluster
                    json_vm_info.get('hostName', ''),  # Host
                    "",  # Folder
                    json_vm_info.get('operatingSystemName', ''),  # OS According to the configuration file
                    vmtoolOS,  # OS according to the VMware Tools
                    json_vm_info.get('moref', ''),  # VM ID
                    json_vm_info.get('id', ''),  # VM UUID
                    json_vm_info.get('machineManagerName', ''),  # VI SDK Server
                    json_vm_info.get('machineManagerId', ''),  # VI SDK UUID
                ]

                write_array_to_row(select_sheet(workbook, "vSnapshot"), snaprowid, vSnap)
                snaprowid = snaprowid + 1

                vSnapHealth = [
                    data_vm['name'],
                    "VM has an active snapshot! " + snap['name'] + ", created on " + str(
                        datetime.fromtimestamp(int(snap['createTime'] / 1000))),
                    "Snapshot",
                    json_vm_info.get('machineManagerName', ''),  # VI SDK Server
                    json_vm_info.get('machineManagerId', ''),  # VI SDK UUID
                ]

                write_array_to_row(select_sheet(workbook, "vHealth"), vhealthrowid, vSnapHealth)
                vhealthrowid = vhealthrowid + 1

            save_workbook(workbook, filepath)
            bar.update(rowid - 1) if PROGRESS else ""

        else:
            log(f"-> Erreur sur la requête pour obtenir les informations de la machine virtuelle avec l'ID : {vm_id} - Code error {request_vm_info.status_code}",
                2)
else:
    log("-> Erreur sur la requête VIRTUAL MACHINES", 1)
    exit()

# Set colomn width based on content and save
for sheetname in workbook.sheetnames:
    dims = {}
    sheet = workbook[sheetname]
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 2))
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value
    sheet.auto_filter.ref = sheet.dimensions
    for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            cell.font = Font(color='FFFFFF')
            cell.fill = blackFill
save_workbook(workbook, filepath)